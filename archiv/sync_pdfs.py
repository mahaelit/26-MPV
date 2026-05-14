"""sync_pdfs.py - Phase 1a des Verifikationsplans.

Nimmt das in Phase 0 bestaetigte PDF->BibKey-Mapping und:

1. kopiert jede Datei aus `MPV/Literatur/` nach
   `MPV/26-MPV/Literatur/<bibkey>/source.<ext>` (ext: pdf oder epub).
2. aktualisiert in `Quellen v0.3.xlsx` fuer jeden gemappten BibKey:
     - Pfad_lokal   = "Literatur/<bibkey>/source.<ext>"
     - Status       = max(aktueller Wert, 3)                  (Volltext im Zugriff)
     - Beschaffung  = nur setzen wenn leer (Default: pdf_online)
     - Bemerkung    = "Nur Inhaltsverzeichnis" fuer die 2 Sonderfaelle anhaengen
   Fuer BibKeys OHNE PDF bleiben die Felder unveraendert.

Idempotent: wiederholte Laeufe ueberschreiben nur was noetig ist.
Die Quellen v0.3.xlsx darf NICHT in Excel geoeffnet sein (Permission Error).
"""
from __future__ import annotations
import shutil
import sys
import unicodedata
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


# ------------- Pfade -------------------------------------------------------
HERE          = Path(__file__).resolve().parent
EXTERNAL_LIT  = HERE.parent / "Literatur"
INTERNAL_LIT  = HERE / "Literatur"
XLSX_PATH     = HERE / "Quellen v0.3.xlsx"


# ------------- Bestaetigtes PDF -> BibKey Mapping (Phase 0) ---------------
# Keys: exakte Dateinamen im externen Literatur-Ordner.
# Values: (BibKey, Notiz-Suffix fuer die Bemerkung)
MAPPING: dict[str, tuple[str, str | None]] = {
    # Volltext-PDFs (17)
    "Utility of Psychometric and Dynamic.pdf":
        ("alhroub2021utility", None),
    "Equitable Identification Practices for.pdf":
        ("alodat2025equitable", None),
    "Migrationsbewegungen und Bevölkerung mit Migrationshintergrund.pdf":
        ("bfs2022migration", None),
    "Hochbegabung_k_ein_Problem.pdf":
        ("brunner2021hochbegabung", None),
    "PISA2022-DieSchweizimFokus.pdf":
        ("erzinger2023pisa", None),
    "Fischer_et_al_2020_Begabungsfoerderung_II.pdf":
        ("fischer2020begabungsfoerderung", None),
    "Promising Practices for Improving Identification of English Learners for.pdf":
        ("gubbins2020promising", None),
    "iPEGE_1_web.pdf":
        ("ipege2009professionelle", None),
    "Begabte_Migration_Infoblatt_I_Keller_110621.pdf":
        ("kellerkoller2011erkennen", None),
    "vonelternmitmigrationshintergrundlernen.pdf":
        ("kosoroklabhart2021voneltern", None),
    "Begabungsförderung und Migrationshintergrund_Dissertation Ulrike Leikhof_Online.pdf":
        ("leikhof2021jugendliche", None),
    "Diagnostik bei Migrantinnen und Migranten _ ein Handbuch  1_ Auflage 2018, Göttingen, 2018 -- Hogrefe Verlag GmbH & Co.pdf":
        ("maehler2018diagnostik", None),
    "Identifying and Serving English Learners in Gifted Education Looking Back and Moving Forward.pdf":
        ("mun2020identifying", None),
    "Hochbegabung _ Erkennen, Verstehen, Fördern -- Preckel, Franzis; Baudson, Tanja Gabriele -- Verlag C_H_ Beck, München, 2013 -- Verlag C_H_.epub":
        ("preckel2013hochbegabung", None),
    "Begabung _ Eine Einführung -- Gabriele Weigand; Victor Müller-Oppliger; Timo Hoyer -- Bookwire GmbH, Darmstadt, 2013 -- WBG Academic --.pdf":
        ("hoyer2013begabung",
         "Einfuehrungsband 2013 (WBG); Schwesterpublikation zum Handbuch Begabung 2021."),
    "HOCHBEGABUNG UND MIGRATIONSHINTERGRUND.pdf":
        ("reutlinger2015hochbegabung", None),
    "Migranten mit Potential Dossier MIRAGE def neu.pdf":
        ("stamm2014mirage", None),
    "Zur_Bedeutung_der_graphomotorischen_Prozesse_beim_.pdf":
        ("sturm2016graphomotorik", None),
    "Uslucan_-_Vortrag_zu_Begabung_-_Daten_und_Fakten.pdf":
        ("uslucan2012begabung", None),

    # Nur-Inhaltsverzeichnis (2) -> Warnung in Bemerkung
    "Inhaltsverzeichnis Alle gleich alle unterschiedlich Zum Umgang mit Heterogenität in Schule und Unterricht.pdf":
        ("buholzer2010allegleich",
         "NUR INHALTSVERZEICHNIS beschafft - Zitate koennen nicht inhaltlich verifiziert werden; bitte Volltext bestellen."),
    "Inhaltsverzeichnis Motivation trifft Begabung Begabte Kinder und Jugendliche verstehen und gezielt fördern.pdf":
        ("lehwald2017motivation",
         "NUR INHALTSVERZEICHNIS beschafft - Zitate koennen nicht inhaltlich verifiziert werden; bitte Volltext bestellen."),
}


def _nfc(s: str) -> str:
    """NFC-Normalisierung: OneDrive speichert Umlaute haeufig als NFD
    (o + Kombinations-Umlaut), unser Literal ist NFC (o-Umlaut). Fuer
    einen verlaesslichen Vergleich normalisieren wir beides auf NFC.
    """
    return unicodedata.normalize("NFC", s)


def _find_file(fname: str) -> Path | None:
    """Findet eine Datei im externen Literatur-Ordner, robust gegenueber
    NFC/NFD-Unicode-Varianten. Gibt den tatsaechlichen Pfad zurueck oder None.
    """
    # Schneller Pfad: exakter Match.
    direct = EXTERNAL_LIT / fname
    if direct.exists():
        return direct
    # Fallback: NFC-normalisierter Name-Vergleich gegen alle Kinder.
    target = _nfc(fname)
    for p in EXTERNAL_LIT.iterdir():
        if _nfc(p.name) == target:
            return p
    return None


def copy_files() -> tuple[set[str], int, list[str]]:
    """Kopiert die Mapping-Dateien in die Zielordner.

    Rueckgabe: (erfolgreich_kopierte_bibkeys, bereits_aktuell, fehlt_liste)
    """
    succeeded: set[str] = set()
    skipped = 0
    missing = []
    for fname, (key, _note) in MAPPING.items():
        src = _find_file(fname)
        if src is None:
            missing.append(fname)
            continue
        ext  = src.suffix.lower().lstrip(".")
        dest_dir = INTERNAL_LIT / key
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / f"source.{ext}"

        need_copy = True
        if dest.exists() and dest.stat().st_size == src.stat().st_size:
            need_copy = False

        if need_copy:
            shutil.copy2(src, dest)
            print(f"  [copy] {src.name[:55]:<55} -> {dest.relative_to(HERE)}")
        else:
            skipped += 1
        succeeded.add(key)
    return succeeded, skipped, missing


def update_xlsx(allowed_keys: set[str]) -> int:
    """Aktualisiert Pfad_lokal / Status / Beschaffung / Bemerkung in XLSX.

    Nur BibKeys, deren Datei tatsaechlich kopiert wurde (`allowed_keys`),
    werden angefasst. Damit wird der XLSX-Zustand niemals vor dem
    Dateisystem-Zustand aktualisiert.

    Rueckgabe: Anzahl aktualisierter Zeilen.
    """
    if not XLSX_PATH.exists():
        print(f"  !! XLSX nicht gefunden: {XLSX_PATH}")
        return 0
    if not allowed_keys:
        print("  (keine BibKeys zu aktualisieren)")
        return 0

    from openpyxl import load_workbook
    try:
        wb = load_workbook(XLSX_PATH)
    except PermissionError:
        print(f"  !! XLSX ist in Excel geoeffnet: {XLSX_PATH}")
        print(f"  !! Bitte schliessen und erneut starten.")
        sys.exit(2)

    ws  = wb["Quellen"] if "Quellen" in wb.sheetnames else wb.active
    hdr = [c.value for c in ws[1]]

    def col_idx(name: str) -> int | None:
        try:
            return hdr.index(name) + 1
        except ValueError:
            return None

    bibkey_i     = col_idx("BibKey")
    pfad_i       = col_idx("Pfad_lokal")
    status_i     = col_idx("Status")
    besch_i      = col_idx("Beschaffung")
    bemerk_i     = col_idx("Bemerkung")

    if bibkey_i is None:
        print("  !! 'BibKey'-Spalte nicht gefunden.")
        sys.exit(2)

    # key -> (ext, note) - aus dem Mapping, fuer schnellen Zugriff.
    by_key: dict[str, tuple[str, str | None]] = {}
    for fname, (key, note) in MAPPING.items():
        src = _find_file(fname)
        ext = (src.suffix if src else ".pdf").lower().lstrip(".") or "pdf"
        by_key[key] = (ext, note)

    updated = 0
    for row in ws.iter_rows(min_row=2):
        key_cell = row[bibkey_i - 1]
        key = key_cell.value
        if not key or key not in allowed_keys:
            continue
        ext, note = by_key[key]
        new_path = f"Literatur/{key}/source.{ext}"

        changed = False

        if pfad_i is not None:
            cell = row[pfad_i - 1]
            if cell.value != new_path:
                cell.value = new_path
                changed = True

        if status_i is not None:
            cell = row[status_i - 1]
            # Status auf mindestens 3 anheben (Volltext vorhanden).
            current_status = cell.value
            try:
                current_int = int(current_status) if current_status not in (None, "") else 0
            except (TypeError, ValueError):
                current_int = 0
            if current_int < 3:
                cell.value = 3
                changed = True

        if besch_i is not None:
            cell = row[besch_i - 1]
            # Nur setzen wenn leer, sonst User-Wert respektieren.
            if cell.value in (None, ""):
                cell.value = "pdf_online"
                changed = True

        if bemerk_i is not None and note:
            cell = row[bemerk_i - 1]
            existing = cell.value or ""
            if note not in existing:
                cell.value = (existing.rstrip() + ("\n" if existing else "") + note).strip()
                changed = True

        if changed:
            updated += 1
            print(f"  [xlsx] {key:35} -> Pfad_lokal + Status=3"
                  + (f" + Beschaffung + Notiz" if note else ""))

    wb.save(XLSX_PATH)
    return updated


def main() -> int:
    print(f"sync_pdfs.py - Phase 1a")
    print(f"  Quelle : {EXTERNAL_LIT}")
    print(f"  Ziel   : {INTERNAL_LIT}")
    print(f"  XLSX   : {XLSX_PATH.name}")
    print()

    print("--- Schritt 1: Dateien kopieren ---")
    succeeded, skipped, missing = copy_files()
    print(f"  Ergebnis: {len(succeeded)} gesamt verfuegbar "
          f"({len(succeeded) - skipped} neu kopiert, {skipped} unveraendert), "
          f"{len(missing)} nicht gefunden.")
    if missing:
        print("  !! Folgende Dateien fehlen im externen Literatur-Ordner:")
        for f in missing:
            print(f"     - {f}")
    print()

    print("--- Schritt 2: XLSX aktualisieren (nur fuer erfolgreich kopierte PDFs) ---")
    updated = update_xlsx(succeeded)
    print(f"  Ergebnis: {updated} Zeilen aktualisiert.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
