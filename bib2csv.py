#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
bib2csv.py  --  Erstellt aus einer BibLaTeX-Datei (z. B. Quellen.bib)
                eine Excel-taugliche CSV- sowie eine XLSX-Datei
                mit einer Zeile pro Literaturquelle.

Aufruf:
    python bib2csv.py [input.bib] [output_basename]

Standardwerte (ohne Argumente):
    input           = Quellen.bib                (neben diesem Skript)
    output_basename = Quellen                    (erzeugt Quellen.csv + Quellen.xlsx)

Besonderheiten
--------------
- Verarbeitet @type{key, feld = {wert}, ...}-Syntax mit
  geschachtelten Klammern und mehrzeiligen Werten.
- Entfernt LaTeX-Markup ({...}, \textit{...}, \,, ~, --, ...)
  fuer eine lesbare Tabelle.
- Erkennt Abschnittsueberschriften der Form
      %  -- Frage 1: Erkennen --
  (Trennzeichen: ASCII '-' oder die Unicode-Box-Zeichen '─' / '═')
  und schreibt den Titel in die Spalte "Kategorie".
- CSV:  UTF-8 mit BOM, Semikolon als Trenner, zusaetzlich Excel-Hinweis
        "sep=;" in Zeile 1, damit die Datei in jeder Excel-Locale mit
        korrekter Spaltentrennung oeffnet.
- XLSX: natives Excel-Format mit fixierter Kopfzeile, fetter Ueberschrift,
        passenden Spaltenbreiten, Textumbruch und klickbaren URL/DOI-Links.
        (Benoetigt das Paket `openpyxl`; faellt bei dessen Fehlen still
        auf CSV-only zurueck.)
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from urllib.parse import quote


# ------------------------------- LaTeX -> Text -------------------------------

_REPLACEMENTS = [
    (re.compile(r"\\textit\{([^{}]*)\}"), r"\1"),
    (re.compile(r"\\emph\{([^{}]*)\}"),   r"\1"),
    (re.compile(r"\\guillemotleft\b"),  "\u00ab"),        # «
    (re.compile(r"\\guillemotright\b"), "\u00bb"),        # »
    (re.compile(r"\\&"), "&"),
    (re.compile(r"\\%"), "%"),
    (re.compile(r"\\,"), " "),
    (re.compile(r"\\ "), " "),
    (re.compile(r"~"),   " "),
    (re.compile(r"--"),  "\u2013"),                      # en-dash
]


def sanitize(value: str) -> str:
    """Entfernt BibTeX-/LaTeX-Steuerzeichen fuer reinen Text in der CSV."""
    if not value:
        return ""
    v = value
    for pat, rep in _REPLACEMENTS:
        v = pat.sub(rep, v)
    # Schutzklammern {...} zweimal entfernen (fuer geschachtelte Paare).
    v = re.sub(r"\{([^{}]*)\}", r"\1", v)
    v = re.sub(r"\{([^{}]*)\}", r"\1", v)
    # Whitespace normalisieren.
    v = re.sub(r"\s+", " ", v).strip()
    return v


def format_names(raw: str) -> str:
    """
    'Last, First and Last2, First2' -> 'Last, First; Last2, First2'.
    Behandelt Umbrueche innerhalb langer Autorlisten korrekt.
    """
    if not raw:
        return ""
    v = re.sub(r"\s+", " ", raw).strip()
    names = [n.strip() for n in v.split(" and ")]
    return "; ".join(sanitize(n) for n in names if n)


# ------------------------------- Bib-Parser ----------------------------------

ENTRY_RE = re.compile(r"@(\w+)\s*\{\s*([^,\s]+)\s*,", re.MULTILINE)

# Abschnittsueberschriften der Form
#     %  --- Titel ---            (ASCII)
#     %  ─── Titel ───            (Unicode Box Drawings)
#     %  ═══ Titel ═══            (Unicode Double Line)
# [ \t] statt \s verhindert, dass die Suche ueber mehrere Zeilen
# hinweg (Leerzeile zwischen Kopf und Titel) zusammenlaeuft.
# Der Titel muss mindestens ein "echtes" Zeichen (kein Trenner,
# kein Whitespace, kein '%') enthalten.
SECTION_RE = re.compile(
    r"^[ \t]*%[ \t]*[\u2500\u2550\-]{2,}[ \t]*"
    r"([^\s%\u2500\u2550\-][^\n]*?)"
    r"[ \t]*[\u2500\u2550\-]{2,}[ \t]*$",
    re.MULTILINE,
)


def find_entry_end(text: str, open_brace: int) -> int:
    """Index der passenden schliessenden `}` zum `{` an *open_brace*."""
    depth = 0
    i = open_brace
    n = len(text)
    while i < n:
        ch = text[i]
        # Kommentar nur AUSSERHALB von Feld-Werten (d. h. direkt innerhalb
        # der Entry-Klammer, depth == 1) als Zeilenkommentar interpretieren.
        if ch == "%" and depth == 1:
            j = text.find("\n", i)
            i = n if j == -1 else j + 1
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return i
        i += 1
    return -1


def strip_comments(text: str) -> str:
    """Entfernt `% ...`-Zeilenkommentare (fuer den Entry-Rumpf)."""
    out = []
    for line in text.splitlines():
        i = 0
        while i < len(line):
            if line[i] == "%" and (i == 0 or line[i - 1] != "\\"):
                break
            i += 1
        out.append(line[:i])
    return "\n".join(out)


def parse_fields(body: str) -> dict:
    """Zerlegt den Rumpf eines @entry{...} in ein Feld-Dictionary."""
    fields: dict = {}
    i, n = 0, len(body)
    while i < n:
        while i < n and body[i] in " \t\n\r,":
            i += 1
        if i >= n:
            break
        m = re.match(r"([A-Za-z_]+)\s*=\s*", body[i:])
        if not m:
            break
        name = m.group(1).lower()
        i += m.end()
        if i >= n:
            break
        if body[i] == "{":
            depth = 1
            i += 1
            start = i
            while i < n and depth > 0:
                if body[i] == "{":
                    depth += 1
                elif body[i] == "}":
                    depth -= 1
                    if depth == 0:
                        break
                i += 1
            value = body[start:i]
            i += 1
        elif body[i] == '"':
            i += 1
            start = i
            while i < n and body[i] != '"':
                i += 1
            value = body[start:i]
            i += 1
        else:
            start = i
            while i < n and body[i] not in ",\n":
                i += 1
            value = body[start:i].strip()
        fields[name] = value
    return fields


_VERIFY_TAG_RE = re.compile(r"^(VERIFY|TODO)\b[^:]*:\s*(.*)", re.IGNORECASE)


def _comment_lines_before(text: str, pos: int) -> list:
    """Sammelt zusammenhaengende '%'-Kommentarzeilen direkt vor *pos*.

    Abbruch bei Leerzeile oder Nicht-Kommentarzeile.
    """
    lines: list = []
    line_start = text.rfind("\n", 0, pos) + 1
    while line_start > 0:
        prev_end = line_start - 1
        if prev_end < 0:
            break
        prev_start = text.rfind("\n", 0, prev_end) + 1
        prev = text[prev_start:prev_end].strip()
        if not prev:
            break                   # Leerzeile  ->  Ende des Blocks
        if not prev.startswith("%"):
            break                   # Code  ->  Ende
        lines.insert(0, prev[1:].strip())
        line_start = prev_start
    return lines


def _collect_verify_todo(comment_lines: list) -> list:
    """Extrahiert VERIFY-/TODO-Notizen (mehrzeilenfaehig) aus Kommentarzeilen."""
    notes: list = []
    current: str | None = None
    for raw in comment_lines:
        line = raw.strip()
        m = _VERIFY_TAG_RE.match(line)
        if m:
            if current:
                notes.append(current)
            tag  = m.group(1).upper()
            body = m.group(2).strip()
            current = f"[{tag}] {body}" if body else f"[{tag}]"
        elif current is not None and line:
            current += " " + line
    if current:
        notes.append(current)
    return notes


def parse_bib(text: str) -> list:
    # Abschnittskoepfe mit ihren Positionen im Originaltext sammeln.
    sections = sorted(
        (m.start(), m.group(1).strip()) for m in SECTION_RE.finditer(text)
    )

    def section_for(pos: int) -> str:
        current = ""
        for sp, name in sections:
            if sp <= pos:
                current = name
            else:
                break
        return current

    entries = []
    for m in ENTRY_RE.finditer(text):
        entry_type = m.group(1).lower()
        key        = m.group(2)
        open_brace = text.find("{", m.start())
        end        = find_entry_end(text, open_brace)
        if end == -1:
            continue
        body_raw   = text[m.end():end]
        body_clean = strip_comments(body_raw)
        fields     = parse_fields(body_clean)

        # VERIFY/TODO-Notizen einsammeln: sowohl aus dem Kommentarblock
        # unmittelbar vor dem Eintrag als auch aus Zeilen innerhalb.
        pre_lines    = _comment_lines_before(text, m.start())
        inline_lines = [
            ln.strip()[1:].strip()
            for ln in body_raw.splitlines()
            if ln.strip().startswith("%")
        ]
        notes = _collect_verify_todo(pre_lines) + _collect_verify_todo(inline_lines)

        fields["entry_type"]  = entry_type
        fields["key"]         = key
        fields["section"]     = section_for(m.start())
        fields["verify_todo"] = " | ".join(notes)
        entries.append(fields)
    return entries


# ------------------------------- CSV-Export ----------------------------------

# Spalten, die aus der .bib-Datei automatisch erzeugt werden.
CSV_COLUMNS = [
    "Kategorie",
    "BibKey",
    "Eintragstyp",
    "AutorIn",
    "HerausgeberIn",
    "Jahr",
    "Titel",
    "Buchtitel / Journal",
    "Auflage",
    "Band",
    "Heft",
    "Seiten",
    "Verlag",
    "Ort",
    "Institution",
    "Schriftenart",
    "Reihe",
    "DOI",
    "URL",
    "Zugriffsdatum",
    "Anmerkung",
    "Umfangshinweis",
]

# ------------------------------- Tracking-Modell -----------------------------
#
# Tracking-Spalten dokumentieren die "externe" Konsistenz jeder Quelle:
# Existenzbestaetigung (Swisscovery/DOI), Volltext-Beschaffung und
# Zitat-gegen-Seite-Verifikation. Sie werden am rechten Rand der Tabelle
# angefuegt, in der XLSX gelb markiert und beim Re-Lauf ueber den "BibKey"
# gemergt. Zusaetzlich werden ALLE weiteren Spalten, die sich in einer
# bestehenden XLSX befinden (z. B. eigene Zusatzspalten), unveraendert
# weitergefuehrt.

TRACKING_COLUMNS = [
    "Status",            # 0-5, siehe STATUS_LEGENDE
    "Swisscovery_URL",   # Permalink zum SLSP-Datensatz (manuell)
    "ISBN_ISSN_DOI",     # Hauptidentifikator (aus .bib vorbefuellt, wenn DOI)
    "Suche_URL",         # Klick-Starter: doi.org / Wayback / Swisscovery-Suche
    "Beschaffung",       # Enum, siehe BESCHAFFUNG_OPTIONS
    "Pfad_lokal",        # relativer Pfad zum Volltext, z. B. Literatur/<key>/source.pdf
    "Verifiziert_am",    # Datum, wann Zitate gegen Quelle gegengeprueft wurden
    "Bemerkung",         # Freitext; VERIFY/TODO-Hinweise aus .bib vorbefuellt
]

# Rueckwaertskompatibel: einige interne Funktionen referenzieren noch
# MANUAL_COLUMNS. Alias auf die Tracking-Spalten.
MANUAL_COLUMNS = TRACKING_COLUMNS

STATUS_OPTIONS = ["0", "1", "2", "3", "4", "5"]

STATUS_LEGENDE = [
    ("0", "ungeprueft",            "Startzustand"),
    ("1", "metadaten_ok",          "Swisscovery-Treffer bestaetigt; Autor/Jahr/Auflage stimmen"),
    ("2", "beschaffung_in_arbeit", "bestellt oder Fernleihe laeuft"),
    ("3", "volltext_lokal",        "PDF/Scan liegt unter Pfad_lokal"),
    ("4", "zitate_verifiziert",    "jede \\cite-Stelle gegen Seite gegengeprueft"),
    ("5", "freigegeben",           "bereit fuer Abgabe"),
]

BESCHAFFUNG_OPTIONS = [
    "ebook_slsp",
    "ebook_verlag",
    "physisch_ausleihe",
    "scan_selbst",
    "pdf_online",
    "fernleihe",
    "nicht_verfuegbar",
]


def _swisscovery_query_parts(entry: dict) -> tuple:
    """Gibt (Namen, Titel) fuer die Swisscovery-Suche zurueck.

    Swisscovery katalogisiert Buecher / Zeitschriften, aber keine
    einzelnen Buchkapitel. Deshalb:

    - @incollection  ->  Suche nach dem ELTERN-Buch:
                         Herausgeber + booktitle
    - alles andere   ->  Autor/Herausgeber + title des Eintrags selbst
    """
    entry_type = (entry.get("entry_type") or "").lower()
    if entry_type == "incollection":
        names_raw = entry.get("editor") or entry.get("author") or ""
        title_raw = entry.get("booktitle") or entry.get("title") or ""
    else:
        names_raw = entry.get("author") or entry.get("editor") or ""
        title_raw = entry.get("title") or ""
    names = format_names(names_raw)
    title = sanitize(title_raw)
    return names, title


def build_search_url(entry: dict) -> str:
    """Generiert eine Verifikations-URL:

    - DOI vorhanden        ->  https://doi.org/<doi>
    - @online mit URL      ->  Wayback-Machine-Snapshot-Uebersicht
    - @incollection        ->  Swisscovery (Herausgeber + Buchtitel)
    - @book/@thesis/...    ->  Swisscovery (Autor/Herausgeber + Titel)
    """
    entry_type = (entry.get("entry_type") or "").lower()
    doi        = sanitize(entry.get("doi", ""))
    url        = sanitize(entry.get("url", ""))

    if doi:
        return f"https://doi.org/{doi}"

    if entry_type == "online" and url:
        return f"https://web.archive.org/web/*/{url}"

    names, title = _swisscovery_query_parts(entry)
    query = " ".join(p for p in (names, title) if p).strip()

    # Kommas und Semikolons bleiben unkodiert (wie im Swisscovery-UI);
    # Leerzeichen werden zu %20, Umlaute zu %XX.
    encoded = quote(query, safe=",;")

    return (
        "https://swisscovery.slsp.ch/discovery/search"
        f"?query=any,contains,{encoded}"
        "&tab=41SLSP_NETWORK"
        "&search_scope=DN_and_CI"
        "&vid=41SLSP_NETWORK:VU1_UNION"
        "&offset=0"
    )


def suggest_local_path(bibkey: str) -> str:
    """Vorschlag fuer Pfad_lokal  ->  'Literatur/<bibkey>/source.pdf'."""
    return f"Literatur/{bibkey}/source.pdf" if bibkey else ""


def tracking_defaults_for(entry: dict) -> dict:
    """Voreinstellungen der Tracking-Spalten aus einer .bib-Entry.

    Diese Defaults werden beim Merge nur dann geschrieben, wenn die
    entsprechende Zelle im bestehenden File leer ist.
    """
    key  = entry.get("key", "")
    doi  = sanitize(entry.get("doi", ""))
    return {
        "Status":          "0",
        "Swisscovery_URL": "",
        "ISBN_ISSN_DOI":   f"doi:{doi}" if doi else "",
        "Suche_URL":       build_search_url(entry),
        "Beschaffung":     "",
        "Pfad_lokal":      suggest_local_path(key),
        "Verifiziert_am":  "",
        "Bemerkung":       entry.get("verify_todo", "") or "",
    }


def entry_to_row(e: dict) -> dict:
    return {
        "Kategorie":            sanitize(e.get("section", "")),
        "BibKey":               e.get("key", ""),
        "Eintragstyp":          e.get("entry_type", ""),
        "AutorIn":              format_names(e.get("author", "")),
        "HerausgeberIn":        format_names(e.get("editor", "")),
        "Jahr":                 sanitize(e.get("year", "")),
        "Titel":                sanitize(e.get("title", "")),
        "Buchtitel / Journal":  sanitize(e.get("booktitle", "") or e.get("journal", "")),
        "Auflage":              sanitize(e.get("edition", "")),
        "Band":                 sanitize(e.get("volume", "")),
        "Heft":                 sanitize(e.get("number", "")),
        "Seiten":               sanitize(e.get("pages", "")),
        "Verlag":               sanitize(e.get("publisher", "")),
        "Ort":                  sanitize(e.get("address", "")),
        "Institution":          sanitize(e.get("institution", "")),
        "Schriftenart":         sanitize(e.get("type", "")),
        "Reihe":                sanitize(e.get("series", "")),
        "DOI":                  sanitize(e.get("doi", "")),
        "URL":                  sanitize(e.get("url", "")),
        "Zugriffsdatum":        sanitize(e.get("urldate", "")),
        "Anmerkung":            sanitize(e.get("note", "")),
        "Umfangshinweis":       sanitize(e.get("annotation", "")),
    }


# ------------------------------- Ausgabeformate ------------------------------

def _looks_like_url(value: str) -> bool:
    return isinstance(value, str) and value.strip().lower().startswith(("http://", "https://"))


def read_manual_data(xlsx_path: Path) -> tuple:
    """Liest manuelle Spalten aus einer bereits vorhandenen XLSX.

    Rueckgabe: (manual_cols, values)
      manual_cols : geordnete Liste der manuellen Spaltennamen
      values      : { BibKey: { Spaltenname: Wert } }
    """
    if not xlsx_path.exists():
        return list(MANUAL_COLUMNS), {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return list(MANUAL_COLUMNS), {}

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as exc:                           # defekte/fremde Datei
        print(f"[Info] Vorhandene XLSX nicht lesbar ({exc})  ->  neu aufbauen.",
              file=sys.stderr)
        return list(MANUAL_COLUMNS), {}

    ws = wb.active
    headers = [c.value for c in ws[1]]
    if "BibKey" not in headers:
        return list(MANUAL_COLUMNS), {}

    bibkey_idx  = headers.index("BibKey")
    auto_cols   = set(CSV_COLUMNS)
    existing_manual: list = []
    manual_idx:      list = []
    for i, h in enumerate(headers):
        if h and h not in auto_cols:
            existing_manual.append(h)
            manual_idx.append((i, h))

    values: dict = {}
    used_columns: set = set()          # Spalten, in denen mind. ein Wert steckt
    for row in ws.iter_rows(min_row=2, values_only=True):
        if bibkey_idx >= len(row):
            continue
        key = row[bibkey_idx]
        if not key:
            continue
        entry = {}
        for i, col_name in manual_idx:
            val = row[i] if i < len(row) else None
            if val not in (None, ""):
                entry[col_name] = val
                used_columns.add(col_name)
        if entry:
            values[str(key)] = entry

    # Reihenfolge: Tracking-Spalten zuerst, dann alle weiteren vorhandenen
    # Spalten -- aber nur jene, die tatsaechlich mindestens einen Wert
    # enthalten. Komplett leere Legacy-Spalten werden stillschweigend
    # verworfen (z. B. "Ebook / Link" aus einer frueheren Version).
    manual_cols = list(TRACKING_COLUMNS)
    for col in existing_manual:
        if col in manual_cols:
            continue
        if col in used_columns:
            manual_cols.append(col)
    return manual_cols, values


# Tracking-Spalten, deren Wert immer aus der aktuellen .bib-Datei neu
# erzeugt wird -- bestehende Zelle wird ueberschrieben. Das ist richtig
# fuer abgeleitete Klick-Starter (z. B. Suche_URL), wo der Nutzer keinen
# stabilen Wert einpflegt. Manuell gepflegte Spalten (Swisscovery_URL,
# Status, etc.) sind NICHT enthalten -- die bleiben erhalten.
ALWAYS_REFRESH = {"Suche_URL"}


def _merge_cell(col_name: str, existing, default: str) -> str:
    """Zellen-Merge.

    - Ist *col_name* in ALWAYS_REFRESH und existiert ein Default, gewinnt
      der Default (die Spalte wird regeneriert).
    - Sonst: bestehender Wert gewinnt, Auto-Default dient als Fallback
      fuer leere Zellen.
    """
    if col_name in ALWAYS_REFRESH and default:
        return default
    if existing not in (None, ""):
        return existing
    return default or ""


def write_csv(rows: list, csv_path: Path,
              manual_cols: list, manual_values: dict,
              tracking_defaults: dict) -> None:
    """Excel-tauglich: UTF-8 mit BOM, ';' als Trenner, 'sep=;'-Praeambel."""
    all_cols = CSV_COLUMNS + list(manual_cols)

    # Bestehende CSV auswerten, damit Werte auch dann erhalten bleiben,
    # wenn aktuell nur eine CSV (keine XLSX) existiert.
    csv_manual = _read_manual_from_csv(csv_path)
    merged_values = {**csv_manual, **manual_values}   # XLSX hat Vorrang

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        # Excel-Hinweis: forciert ';' als Spaltentrenner, unabhaengig von
        # der System-Locale (auch in EN/US-Excel funktioniert die Datei).
        f.write("sep=;\n")
        writer = csv.DictWriter(f, fieldnames=all_cols, delimiter=";")
        writer.writeheader()
        for row in rows:
            key      = row.get("BibKey", "")
            manual   = merged_values.get(key, {})
            defaults = tracking_defaults.get(key, {})
            merged   = {c: _merge_cell(c, manual.get(c, ""), defaults.get(c, ""))
                        for c in manual_cols}
            writer.writerow({**row, **merged})


def _read_manual_from_csv(csv_path: Path) -> dict:
    """Fallback-Merge aus CSV (verwendet, wenn keine XLSX existiert)."""
    if not csv_path.exists():
        return {}
    try:
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            first = f.readline()
            if not first.lower().startswith("sep="):
                f.seek(0)
            reader = csv.DictReader(f, delimiter=";")
            auto = set(CSV_COLUMNS)
            manual_cols = [h for h in (reader.fieldnames or []) if h and h not in auto]
            if not manual_cols:
                return {}
            result = {}
            for r in reader:
                key = (r.get("BibKey") or "").strip()
                if not key:
                    continue
                entry = {c: r.get(c, "") for c in manual_cols if r.get(c)}
                if entry:
                    result[key] = entry
            return result
    except Exception:
        return {}


def write_xlsx(rows: list, xlsx_path: Path,
               manual_cols: list, manual_values: dict,
               tracking_defaults: dict) -> bool:
    """Echte .xlsx-Datei schreiben. Gibt False zurueck, falls openpyxl fehlt."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule
    except ImportError:
        print("[Info] openpyxl nicht installiert  ->  nur CSV erzeugt.",
              file=sys.stderr)
        print("       Installieren mit:  pip install openpyxl", file=sys.stderr)
        return False

    all_cols = CSV_COLUMNS + list(manual_cols)
    n_rows   = len(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Quellen"

    # Kopfzeile.
    ws.append(all_cols)
    header_font_auto    = Font(bold=True, color="FFFFFF")
    header_fill_auto    = PatternFill("solid", fgColor="305496")   # dunkelblau
    header_font_track   = Font(bold=True, color="7F6000")
    header_fill_track   = PatternFill("solid", fgColor="FFE699")   # warmes gelb
    header_align        = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx, name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if name in manual_cols:
            cell.font = header_font_track
            cell.fill = header_fill_track
        else:
            cell.font = header_font_auto
            cell.fill = header_fill_auto
        cell.alignment = header_align

    # Hilfsstile.
    link_font    = Font(color="0563C1", underline="single")
    manual_fill  = PatternFill("solid", fgColor="FFF2CC")          # sehr helles gelb
    top_wrap     = Alignment(vertical="top", wrap_text=True)
    url_col      = CSV_COLUMNS.index("URL") + 1
    doi_col      = CSV_COLUMNS.index("DOI") + 1
    manual_col_indexes = {c: all_cols.index(c) + 1 for c in manual_cols}

    for r_idx, row in enumerate(rows, start=2):
        key      = row.get("BibKey", "")
        manual   = manual_values.get(key, {})
        defaults = tracking_defaults.get(key, {})

        # Auto-Spalten.
        for c_idx, col_name in enumerate(CSV_COLUMNS, start=1):
            value = row.get(col_name, "") or ""
            cell  = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = top_wrap

        # Manuelle / Tracking-Spalten.
        for col_name, c_idx in manual_col_indexes.items():
            existing = manual.get(col_name, "")
            default  = defaults.get(col_name, "")
            value    = _merge_cell(col_name, existing, default)
            cell = ws.cell(row=r_idx, column=c_idx, value=value or None)
            cell.alignment = top_wrap
            cell.fill      = manual_fill
            if _looks_like_url(value):
                cell.hyperlink = value
                cell.font      = link_font

        # URL als Hyperlink.
        url_val = row.get("URL", "")
        if url_val:
            cell = ws.cell(row=r_idx, column=url_col)
            cell.hyperlink = url_val
            cell.font      = link_font

        # DOI als Hyperlink.
        doi_val = row.get("DOI", "")
        if doi_val:
            cell = ws.cell(row=r_idx, column=doi_col)
            cell.hyperlink = f"https://doi.org/{doi_val}"
            cell.font      = link_font

    # Kopfzeile fixieren + Autofilter.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Spaltenbreiten.
    max_widths = {
        "Kategorie":            28,
        "BibKey":               28,
        "Eintragstyp":          14,
        "AutorIn":              40,
        "HerausgeberIn":        40,
        "Jahr":                 6,
        "Titel":                55,
        "Buchtitel / Journal":  45,
        "Auflage":              8,
        "Band":                 6,
        "Heft":                 6,
        "Seiten":               12,
        "Verlag":               30,
        "Ort":                  20,
        "Institution":          30,
        "Schriftenart":         16,
        "Reihe":                20,
        "DOI":                  30,
        "URL":                  45,
        "Zugriffsdatum":        14,
        "Anmerkung":            40,
        "Umfangshinweis":       40,
        "Status":               7,
        "Swisscovery_URL":      45,
        "ISBN_ISSN_DOI":        26,
        "Suche_URL":            45,
        "Beschaffung":          22,
        "Pfad_lokal":           45,
        "Verifiziert_am":       14,
        "Bemerkung":            45,
    }
    for idx, col_name in enumerate(all_cols, start=1):
        if col_name in CSV_COLUMNS:
            content_w = max(
                (len(str(row.get(col_name, "") or "")) for row in rows),
                default=0,
            )
        else:
            values = []
            for r in rows:
                key      = r.get("BibKey", "")
                existing = manual_values.get(key, {}).get(col_name, "")
                default  = tracking_defaults.get(key, {}).get(col_name, "")
                values.append(str(_merge_cell(col_name, existing, default) or ""))
            content_w = max((len(v) for v in values), default=0)
        header_w = len(col_name)
        cap      = max_widths.get(col_name, 35)
        width    = min(cap, max(header_w + 2, min(content_w + 2, 60)))
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 28

    # --- Excel-Komfort: Dropdown-Validierung fuer Enum-Spalten --------------
    if n_rows > 0 and "Status" in manual_col_indexes:
        col_letter = get_column_letter(manual_col_indexes["Status"])
        status_dv = DataValidation(
            type="list",
            formula1='"' + ",".join(STATUS_OPTIONS) + '"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Ungueltiger Status",
            error="Bitte einen Wert von 0 bis 5 waehlen (siehe Legende).",
            promptTitle="Status",
            prompt="0=ungeprueft, 1=metadaten_ok, 2=beschaffung, 3=volltext, 4=zitate, 5=freigegeben",
        )
        status_dv.add(f"{col_letter}2:{col_letter}{n_rows + 1}")
        ws.add_data_validation(status_dv)

    if n_rows > 0 and "Beschaffung" in manual_col_indexes:
        col_letter = get_column_letter(manual_col_indexes["Beschaffung"])
        besch_dv = DataValidation(
            type="list",
            formula1='"' + ",".join(BESCHAFFUNG_OPTIONS) + '"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Ungueltige Beschaffung",
            error="Bitte eine der vordefinierten Optionen waehlen.",
            promptTitle="Beschaffungsart",
            prompt="ebook_slsp / ebook_verlag / physisch_ausleihe / scan_selbst / pdf_online / fernleihe / nicht_verfuegbar",
        )
        besch_dv.add(f"{col_letter}2:{col_letter}{n_rows + 1}")
        ws.add_data_validation(besch_dv)

    # --- Ampel-Formatierung fuer Status (0 rot, 1-2 orange, 3 gelb, 4-5 gruen) ---
    if n_rows > 0 and "Status" in manual_col_indexes:
        col_letter = get_column_letter(manual_col_indexes["Status"])
        rng        = f"{col_letter}2:{col_letter}{n_rows + 1}"
        rules = [
            (["0"],        "F4CCCC"),   # rot
            (["1", "2"],   "FCE5CD"),   # orange
            (["3"],        "FFF2CC"),   # gelb
            (["4", "5"],   "D9EAD3"),   # gruen
        ]
        for values, color in rules:
            for v in values:
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator="equal",
                        formula=[f'"{v}"'],
                        fill=PatternFill(start_color=color,
                                         end_color=color,
                                         fill_type="solid"),
                    ),
                )

    # --- Legende-Tab -----------------------------------------------------------
    legende = wb.create_sheet("Legende")
    legende.append(["Status", "Bezeichnung", "Bedeutung"])
    for c in legende[1]:
        c.font = header_font_auto
        c.fill = header_fill_auto
        c.alignment = header_align
    for status, name, desc in STATUS_LEGENDE:
        legende.append([status, name, desc])
    legende.column_dimensions["A"].width = 8
    legende.column_dimensions["B"].width = 24
    legende.column_dimensions["C"].width = 80
    legende.append([])
    legende.append(["Beschaffung", "Bedeutung"])
    for c in legende[legende.max_row]:
        c.font = header_font_auto
        c.fill = header_fill_auto
        c.alignment = header_align
    beschaffung_erklaerung = {
        "ebook_slsp":         "E-Book ueber SLSP-Lizenz (PH-Luzern-Login)",
        "ebook_verlag":       "E-Book direkt beim Verlag gekauft/geliehen",
        "physisch_ausleihe":  "Printexemplar via Swisscovery-Kurier an PH-Luzern / ZHB",
        "scan_selbst":        "relevantes Kapitel nach Ausleihe selbst gescannt (URG Art. 19)",
        "pdf_online":         "Open-Access-PDF direkt heruntergeladen",
        "fernleihe":          "Fernleihe ausserhalb SLSP (z. B. KVK/Deutschland)",
        "nicht_verfuegbar":   "derzeit nicht beschaffbar, Alternative pruefen",
    }
    for opt in BESCHAFFUNG_OPTIONS:
        legende.append([opt, beschaffung_erklaerung.get(opt, "")])

    wb.save(xlsx_path)
    return True


# ------------------------------- Ordner-Scaffolding --------------------------

_VERIFIED_QUOTES_TEMPLATE = """# Verifizierte Zitate – {bibkey}

**Quelle:** {author}{year}. {title}.{venue}
**Swisscovery/Verifikationslink:** {search_url}
**Identifikator:** {ident}
**Lokaler Pfad:** `source.pdf`

---

## Zitate (gegen die Quelle gegengeprueft)

### Zitat 1 (S. XX)

> „Wortgetreues Zitat hier einfuegen.“

**Kontext / Paraphrase:**
<eigene Zusammenfassung in 1-2 Saetzen>

**Verwendet in:**
- Lerndokument: §<Abschnitt>
- Abgabedokument: §<Abschnitt>

---

### Zitat 2 (S. XX)

> „…“

<...>

---

**Status:** 0 (ungeprueft)
**Verifiziert am:** <YYYY-MM-DD>
**Bearbeitet durch:** Inti Merolli
"""


def _render_verified_quotes(entry: dict) -> str:
    """Baut den vollstaendigen Inhalt einer verified_quotes.md-Vorlage."""
    key    = entry.get("key", "")
    author = format_names(entry.get("author", "") or entry.get("editor", "")) or "(Autor unbekannt)"
    year   = sanitize(entry.get("year", ""))
    title  = sanitize(entry.get("title", ""))
    venue  = ""
    bt     = sanitize(entry.get("booktitle", ""))
    jn     = sanitize(entry.get("journal", ""))
    if bt:
        venue = f" In: {bt}."
    elif jn:
        venue = f" {jn}."
    ident = sanitize(entry.get("doi", ""))
    ident = f"doi:{ident}" if ident else "–"
    return _VERIFIED_QUOTES_TEMPLATE.format(
        bibkey     = key,
        author     = author,
        year       = f" ({year})" if year else "",
        title      = title or "(kein Titel)",
        venue      = venue,
        search_url = build_search_url(entry),
        ident      = ident,
    )


_MD_HEADER_SEP_RE = re.compile(r"\n\s*---\s*\n")


def _refresh_md_header(existing: str, expected: str) -> str | None:
    """Ersetzt den Metadaten-Header (bis zum ersten '---'-Trenner) mit dem
    frisch generierten Header und laesst den Rest (User-Notizen) intakt.

    Gibt None zurueck, wenn das Format nicht erkannt wird.
    """
    me = _MD_HEADER_SEP_RE.search(existing)
    mx = _MD_HEADER_SEP_RE.search(expected)
    if not (me and mx):
        return None
    new_header = expected[:mx.end()]
    old_body   = existing[me.end():]
    return new_header + old_body


def scaffold_literatur_folders(entries: list, base_dir: Path) -> tuple:
    """Legt fuer jede Quelle einen Ordner `Literatur/<BibKey>/` an und
    schreibt eine Vorlage `verified_quotes.md` hinein.

    - Neue Ordner/Vorlagen werden angelegt.
    - Bestehende verified_quotes.md-Dateien bekommen nur den Metadaten-
      **Header** (bis zum ersten `---`-Trenner) aufgefrischt; Zitate und
      Notizen darunter bleiben unveraendert.

    Rueckgabe: (neue_ordner, neue_vorlagen, aufgefrischte_header)
    """
    lit_root = base_dir / "Literatur"
    lit_root.mkdir(exist_ok=True)

    new_dirs      = 0
    new_files     = 0
    refreshed_hdr = 0
    for e in entries:
        key = e.get("key", "")
        if not key:
            continue
        folder = lit_root / key
        if not folder.exists():
            folder.mkdir(parents=True, exist_ok=True)
            new_dirs += 1

        md_path  = folder / "verified_quotes.md"
        expected = _render_verified_quotes(e)

        if not md_path.exists():
            md_path.write_text(expected, encoding="utf-8")
            new_files += 1
        else:
            existing = md_path.read_text(encoding="utf-8")
            refreshed = _refresh_md_header(existing, expected)
            if refreshed is not None and refreshed != existing:
                md_path.write_text(refreshed, encoding="utf-8")
                refreshed_hdr += 1
    return new_dirs, new_files, refreshed_hdr


# ------------------------------- Haupt-Einstieg ------------------------------

def main(argv) -> int:
    here = Path(__file__).resolve().parent
    bib_path = Path(argv[1]) if len(argv) >= 2 else here / "Quellen.bib"
    if len(argv) >= 3:
        base = Path(argv[2])
        base = base.with_suffix("") if base.suffix.lower() in (".csv", ".xlsx") else base
    else:
        base = here / "Quellen"
    csv_path  = base.with_suffix(".csv")
    xlsx_path = base.with_suffix(".xlsx")

    if not bib_path.is_file():
        print(f"[Fehler] Eingabedatei nicht gefunden: {bib_path}", file=sys.stderr)
        return 1

    text    = bib_path.read_text(encoding="utf-8")
    entries = parse_bib(text)
    rows    = [entry_to_row(e) for e in entries]

    # Auto-Defaults fuer die Tracking-Spalten, pro BibKey.
    tracking_defaults = {e.get("key", ""): tracking_defaults_for(e) for e in entries}

    # Manuelle Werte aus einer bereits vorhandenen XLSX uebernehmen, damit
    # von Hand eingetragene Werte beim Re-Lauf nicht verloren gehen.
    manual_cols, manual_values = read_manual_data(xlsx_path)
    preserved = sum(len(v) for v in manual_values.values())
    if preserved:
        print(f"[OK] Manuelle Werte uebernommen: {preserved} Zellen in "
              f"{len(manual_values)} Zeilen ({', '.join(manual_cols)})")

    try:
        write_csv(rows, csv_path, manual_cols, manual_values, tracking_defaults)
        print(f"[OK] {len(rows):>3} Eintraege  ->  {csv_path}")
    except PermissionError:
        print(f"[Fehler] CSV ist gesperrt (in Excel offen?): {csv_path}",
              file=sys.stderr)
        return 2

    try:
        if write_xlsx(rows, xlsx_path, manual_cols, manual_values, tracking_defaults):
            print(f"[OK] {len(rows):>3} Eintraege  ->  {xlsx_path}")
    except PermissionError:
        print(f"[Fehler] XLSX ist gesperrt (in Excel offen?): {xlsx_path}",
              file=sys.stderr)
        return 2

    # Ordnerstruktur 'Literatur/<BibKey>/' mit verified_quotes.md anlegen
    # bzw. Header bestehender Dateien aktualisieren.
    new_dirs, new_files, refreshed = scaffold_literatur_folders(entries, here)
    parts = []
    if new_dirs:
        parts.append(f"+{new_dirs} Verzeichnisse")
    if new_files:
        parts.append(f"+{new_files} Quote-Vorlagen")
    if refreshed:
        parts.append(f"{refreshed} Header aktualisiert")
    if parts:
        print("[OK] Literaturordner: " + ", ".join(parts) + ".")
    else:
        print("[OK] Literaturordner/Quote-Vorlagen unveraendert.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
