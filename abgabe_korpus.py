#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""Generate the definitive MPV submission source list from LaTeX markers.

The source of truth is the set of \AbgabeQuelle commands reachable from the
configured TeX file. Metadata is resolved from Quellen.bib. The script writes:

- Abgabe_Quellen.csv
- Abgabe_Quellen.xlsx, if openpyxl is installed
- build/abgabe_seitenbudget.tex
- build/abgabe_quellenliste.tex
- build/abgabe_korpus_audit.md
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from bib2csv import entry_to_row, format_names, parse_bib, sanitize


COUNTED_STATUSES = {"K", "S"}
VALID_STATUSES = {"K", "S", "N"}

STATUS_LABELS = {
    "K": "Kernliteratur",
    "S": "gezaehlte Stuetzliteratur",
    "N": "nicht gezaehlt",
}

QUESTION_LABELS = {
    "F1": "Frage 1 (FW, KS/DG: Erkennen)",
    "F2": "Frage 2 (FW, MW/KS: Barrieren)",
    "F3": "Frage 3 (FW, PB: Beziehung)",
    "F4": "Frage 4 (BW, DG/PV: Setting)",
    "F5": "Frage 5 (BW, ZB/PV: Profession)",
}

QUESTION_FOCUS = {
    "F1": "Begabungstheorie, Diagnostik, 2e, Migration",
    "F2": "Graphomotorik, Schriftsprache, GRAFINK",
    "F3": "Anerkennung, Peer-Beziehung, Teilhabe",
    "F4": "Enrichment, Inklusion, Bildungsgerechtigkeit, Schweizer Volksschule",
    "F5": "Beratung, Kooperation, Profession",
}

EXCLUDED_COUNTED_KEYS = {
    "mun2020identifying",
    "gubbins2020promising",
    "alhroub2021utility",
    "alodat2025equitable",
}

SOURCE_RE = re.compile(
    r"\\AbgabeQuelle\s*"
    r"\{(?P<frage>[^{}]+)\}\s*"
    r"\{(?P<status>[^{}]+)\}\s*"
    r"\{(?P<key>[^{}]+)\}\s*"
    r"\{(?P<seiten>[^{}]+)\}\s*"
    r"\{(?P<zahl>[^{}]+)\}",
    re.DOTALL,
)
INPUT_RE = re.compile(r"\\(?:input|include)\{([^{}]+)\}")
UMFANG_RE = re.compile(
    r"Umfang\s+Frage~?([1-5])\s*:\s*([0-9]+)(?:\\,|\s|~)*Seiten",
    re.IGNORECASE,
)
MANUAL_BUDGET_RE = re.compile(
    r"Frage\s+([1-5])\s*\([^&]+&\s*([0-9]+)\s*&",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class SourceRecord:
    index: int
    frage: str
    status: str
    key: str
    seiten: str
    seitenzahl: int

    @property
    def counted(self) -> bool:
        return self.status in COUNTED_STATUSES

    def signature(self) -> tuple[str, str, str, str, int]:
        return (self.frage, self.status, self.key, self.seiten, self.seitenzahl)


class KorpusError(RuntimeError):
    pass


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def collect_tex(path: Path, seen: set[Path] | None = None) -> str:
    r"""Read a TeX file and inline local \input/\include files once."""
    seen = seen or set()
    path = path.resolve()
    if path in seen:
        return ""
    seen.add(path)

    text = read_text(path)

    def replace_input(match: re.Match[str]) -> str:
        raw = match.group(1).strip()
        candidate = Path(raw)
        if not candidate.suffix:
            candidate = candidate.with_suffix(".tex")
        if not candidate.is_absolute():
            candidate = path.parent / candidate
        if not candidate.exists():
            return match.group(0)
        return "\n" + collect_tex(candidate, seen) + "\n"

    return INPUT_RE.sub(replace_input, text)


def strip_latex_comments(text: str) -> str:
    cleaned_lines: list[str] = []
    for line in text.splitlines():
        cut = len(line)
        for idx, char in enumerate(line):
            if char == "%" and (idx == 0 or line[idx - 1] != "\\"):
                cut = idx
                break
        cleaned_lines.append(line[:cut])
    return "\n".join(cleaned_lines)


def parse_sources(tex_text: str) -> list[SourceRecord]:
    tex_text = strip_latex_comments(tex_text)
    records: list[SourceRecord] = []
    for idx, match in enumerate(SOURCE_RE.finditer(tex_text), start=1):
        frage = match.group("frage").strip().upper()
        status = match.group("status").strip().upper()
        key = match.group("key").strip()
        seiten = " ".join(match.group("seiten").split())
        zahl_raw = match.group("zahl").strip()
        if not re.fullmatch(r"[0-9]+", zahl_raw):
            raise KorpusError(
                f"Seitenzahl ist nicht numerisch: {key} ({frage}, {zahl_raw!r})"
            )
        records.append(
            SourceRecord(
                index=idx,
                frage=frage,
                status=status,
                key=key,
                seiten=seiten,
                seitenzahl=int(zahl_raw),
            )
        )
    if not records:
        raise KorpusError("Keine \\AbgabeQuelle-Eintraege im TeX-Korpus gefunden.")
    return records


def load_bib(path: Path) -> dict[str, dict]:
    entries = parse_bib(read_text(path))
    by_key: dict[str, dict] = {}
    duplicates: list[str] = []
    for entry in entries:
        key = entry.get("key", "")
        if key in by_key:
            duplicates.append(key)
        by_key[key] = entry
    if duplicates:
        raise KorpusError("Doppelte BibTeX-Keys in Quellen.bib: " + ", ".join(duplicates))
    return by_key


def totals_by_question(records: Iterable[SourceRecord]) -> dict[str, int]:
    totals = {frage: 0 for frage in QUESTION_LABELS}
    for record in records:
        if record.counted:
            totals[record.frage] += record.seitenzahl
    return totals


def validate_records(records: list[SourceRecord], bib: dict[str, dict], tex_text: str) -> None:
    errors: list[str] = []

    seen_signatures: set[tuple[str, str, str, str, int]] = set()
    for record in records:
        if record.frage not in QUESTION_LABELS:
            errors.append(f"Ungueltige Frage {record.frage!r} bei {record.key}.")
        if record.status not in VALID_STATUSES:
            errors.append(f"Ungueltiger Status {record.status!r} bei {record.key}.")
        if record.key not in bib:
            errors.append(f"BibTeX-Key aus LaTeX fehlt in Quellen.bib: {record.key}.")
        if record.status == "N" and record.seitenzahl != 0:
            errors.append(
                f"Nicht gezaehlte Quelle darf keine Seiten zaehlen: {record.key}."
            )
        if record.status in COUNTED_STATUSES and record.seitenzahl <= 0:
            errors.append(f"Gezaehlte Quelle hat keine positive Seitenzahl: {record.key}.")
        if record.signature() in seen_signatures:
            errors.append(
                "Doppelter Korpuseintrag: "
                f"{record.frage}/{record.status}/{record.key}/{record.seiten}."
            )
        seen_signatures.add(record.signature())

        if record.key in EXCLUDED_COUNTED_KEYS and record.counted:
            errors.append(
                "Ergaenzender Fachartikel versehentlich gezaehlt: "
                f"{record.key} ({record.frage})."
            )
        if record.key in bib and record.counted:
            section = sanitize(bib[record.key].get("section", "")).lower()
            if "fachartikel-exzerpte" in section:
                errors.append(
                    "BibTeX-Abschnitt 'Ergaenzungsliteratur: Fachartikel-Exzerpte' "
                    f"darf nicht gezaehlt werden: {record.key}."
                )

    computed = totals_by_question(records)
    for frage_num, claimed in parse_inline_umfangs(tex_text).items():
        frage = f"F{frage_num}"
        if frage in computed and computed[frage] != claimed:
            errors.append(
                f"Umfang-Angabe im LaTeX passt nicht zu \\AbgabeQuelle: "
                f"{frage} LaTeX={claimed}, Korpus={computed[frage]}."
            )

    if "abgabe_seitenbudget.tex" not in tex_text:
        for frage_num, budget in parse_manual_budget(tex_text).items():
            frage = f"F{frage_num}"
            if frage in computed and computed[frage] != budget:
                errors.append(
                    f"Manuelle Seitenbudget-Tabelle passt nicht zum Korpus: "
                    f"{frage} Tabelle={budget}, Korpus={computed[frage]}."
                )

    if errors:
        raise KorpusError("\n".join(f"- {err}" for err in errors))


def parse_inline_umfangs(tex_text: str) -> dict[str, int]:
    return {m.group(1): int(m.group(2)) for m in UMFANG_RE.finditer(tex_text)}


def parse_manual_budget(tex_text: str) -> dict[str, int]:
    return {m.group(1): int(m.group(2)) for m in MANUAL_BUDGET_RE.finditer(tex_text)}


def canonical_payload(records: list[SourceRecord], bib: dict[str, dict]) -> dict:
    payload_records = []
    for record in sorted(records, key=lambda r: (r.frage, r.status, r.key, r.seiten)):
        entry = bib[record.key]
        row = entry_to_row(entry)
        payload_records.append(
            {
                "frage": record.frage,
                "status": record.status,
                "key": record.key,
                "seiten": record.seiten,
                "seitenzahl": record.seitenzahl,
                "autor": row.get("AutorIn", ""),
                "jahr": row.get("Jahr", ""),
                "titel": row.get("Titel", ""),
                "venue": row.get("Buchtitel / Journal", ""),
            }
        )
    return {"schema": "mpv-abgabe-korpus-v1", "records": payload_records}


def corpus_hash(records: list[SourceRecord], bib: dict[str, dict]) -> str:
    raw = json.dumps(
        canonical_payload(records, bib),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def row_for(record: SourceRecord, bib: dict[str, dict], digest: str) -> dict[str, str | int]:
    entry = bib[record.key]
    bib_row = entry_to_row(entry)
    return {
        "Frage": record.frage,
        "Frage_Label": QUESTION_LABELS[record.frage],
        "Status": record.status,
        "Status_Text": STATUS_LABELS[record.status],
        "Gezaehlt": "ja" if record.counted else "nein",
        "BibKey": record.key,
        "Seitenbereich": record.seiten,
        "Seitenzahl": record.seitenzahl,
        "AutorIn": bib_row.get("AutorIn", ""),
        "HerausgeberIn": bib_row.get("HerausgeberIn", ""),
        "Jahr": bib_row.get("Jahr", ""),
        "Titel": bib_row.get("Titel", ""),
        "Buchtitel / Journal": bib_row.get("Buchtitel / Journal", ""),
        "BibTeX-Seiten": bib_row.get("Seiten", ""),
        "DOI": bib_row.get("DOI", ""),
        "URL": bib_row.get("URL", ""),
        "BibTeX-Annotation": bib_row.get("Umfangshinweis", ""),
        "Korpus_Pruefsumme": digest,
    }


COLUMNS = [
    "Frage",
    "Frage_Label",
    "Status",
    "Status_Text",
    "Gezaehlt",
    "BibKey",
    "Seitenbereich",
    "Seitenzahl",
    "AutorIn",
    "HerausgeberIn",
    "Jahr",
    "Titel",
    "Buchtitel / Journal",
    "BibTeX-Seiten",
    "DOI",
    "URL",
    "BibTeX-Annotation",
    "Korpus_Pruefsumme",
]


def write_csv(rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        handle.write("sep=;\n")
        writer = csv.DictWriter(handle, fieldnames=COLUMNS, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def existing_xlsx_signature(path: Path) -> set[tuple[str, str, str, str, int]] | None:
    if not path.exists():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        raise KorpusError(f"Bestehende XLSX ist nicht lesbar: {path} ({exc})") from exc

    signatures: set[tuple[str, str, str, str, int]] = set()
    expected_sheets = ["Abgabequellen", "Nicht_gezaehlt"]
    if not any(name in wb.sheetnames for name in expected_sheets):
        raise KorpusError(
            f"Bestehende XLSX hat nicht das erwartete Generatorformat: {path}. "
            "Mit --force bewusst neu erzeugen."
        )

    for sheet_name in expected_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        required = ["Frage", "Status", "BibKey", "Seitenbereich", "Seitenzahl"]
        if not all(col in headers for col in required):
            raise KorpusError(
                f"Bestehende XLSX-Sheet {sheet_name!r} hat nicht alle Pflichtspalten."
            )
        idx = {col: headers.index(col) for col in required}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[idx["BibKey"]]:
                continue
            try:
                count = int(row[idx["Seitenzahl"]])
            except (TypeError, ValueError) as exc:
                raise KorpusError(
                    f"Nicht numerische Seitenzahl in bestehender XLSX: {row}"
                ) from exc
            signatures.add(
                (
                    str(row[idx["Frage"]]),
                    str(row[idx["Status"]]),
                    str(row[idx["BibKey"]]),
                    str(row[idx["Seitenbereich"]]),
                    count,
                )
            )
    return signatures


def enforce_existing_xlsx_matches(path: Path, records: list[SourceRecord], force: bool) -> None:
    if force or not path.exists():
        return
    existing = existing_xlsx_signature(path)
    if existing is None:
        return
    expected = {record.signature() for record in records}
    extras = sorted(existing - expected)
    missing = sorted(expected - existing)
    if extras or missing:
        parts = [
            f"Bestehende XLSX stimmt nicht mit dem LaTeX-Korpus ueberein: {path}"
        ]
        if extras:
            parts.append("In Excel, aber nicht im LaTeX-Korpus:")
            parts.extend(f"  - {item}" for item in extras[:20])
        if missing:
            parts.append("Im LaTeX-Korpus, aber nicht in Excel:")
            parts.extend(f"  - {item}" for item in missing[:20])
        parts.append("Nach bewusster Korpus-Aenderung mit --force neu erzeugen.")
        raise KorpusError("\n".join(parts))


def write_xlsx(
    rows_counted: list[dict],
    rows_uncounted: list[dict],
    totals: dict[str, int],
    digest: str,
    path: Path,
    tex_path: Path,
    bib_path: Path,
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Info] openpyxl fehlt; XLSX wird nicht erzeugt.", file=sys.stderr)
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Abgabequellen"
    add_rows_sheet(ws, rows_counted)

    ws_n = wb.create_sheet("Nicht_gezaehlt")
    add_rows_sheet(ws_n, rows_uncounted)

    ws_b = wb.create_sheet("Seitenbudget")
    ws_b.append(["Frage", "Seiten", "Schwerpunkt"])
    for frage in QUESTION_LABELS:
        ws_b.append([QUESTION_LABELS[frage], totals[frage], QUESTION_FOCUS[frage]])
    ws_b.append(["Summe", sum(totals.values()), ""])

    ws_m = wb.create_sheet("Metadaten")
    ws_m.append(["Feld", "Wert"])
    ws_m.append(["Korpus_Pruefsumme", digest])
    ws_m.append(["Quelle_TeX", str(tex_path)])
    ws_m.append(["Quelle_Bib", str(bib_path)])
    ws_m.append(["Erzeugt_UTC", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    ws_m.append(["Generator", Path(__file__).name])

    for sheet in wb.worksheets:
        style_sheet(sheet)

    wb.save(path)
    return True


def add_rows_sheet(ws, rows: list[dict]) -> None:
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])


def style_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    top_wrap = Alignment(vertical="top", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = top_wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    headers = [cell.value for cell in ws[1]]
    url_columns = {headers.index(col) + 1 for col in ("URL", "DOI") if col in headers}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = top_wrap
            if cell.column in url_columns and cell.value:
                value = str(cell.value)
                if cell.column == headers.index("DOI") + 1 and not value.startswith("http"):
                    cell.hyperlink = f"https://doi.org/{value}"
                elif value.startswith(("http://", "https://")):
                    cell.hyperlink = value
                if cell.hyperlink:
                    cell.font = link_font

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header or ""))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 70))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 55)


def tex_escape_text(value: str) -> str:
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
    }
    return "".join(replacements.get(ch, ch) for ch in value)


def write_budget_tex(totals: dict[str, int], digest: str, path: Path) -> None:
    total = sum(totals.values())
    lines = [
        "% Diese Datei wird von abgabe_korpus.py erzeugt. Nicht manuell bearbeiten.",
        f"% Korpus-Pruefsumme: {digest}",
        r"\begin{tabularx}{\textwidth}{Xr@{\hspace{1.5em}}X}",
        r"\toprule",
        r"\textbf{Frage} & \textbf{S.} & \textbf{Schwerpunkt} \\",
        r"\midrule",
    ]
    for frage in QUESTION_LABELS:
        lines.append(
            f"{QUESTION_LABELS[frage]} & {totals[frage]} & "
            f"{tex_escape_text(QUESTION_FOCUS[frage])} " + r"\\"
        )
    lines.extend(
        [
            r"\midrule",
            rf"\textbf{{Summe fragespezifisch}} & \textbf{{{total}}} & \\",
            r"\bottomrule",
            r"\end{tabularx}",
            "",
            r"\vspace{0.8em}",
            r"\textit{Diese Tabelle wurde automatisch aus "
            r"\texttt{abgabe\_korpus.tex} und \texttt{Quellen.bib} erzeugt. "
            r"Gezaehlt werden ausschliesslich Eintraege mit Status "
            r"\texttt{K} oder \texttt{S}; nicht gezaehlte Stuetzliteratur "
            r"und ergaenzende Fachartikel bleiben ausserhalb der Summe. "
            rf"Korpus-Pruefsumme: \texttt{{{digest[:16]}}}.}}",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_quellenliste_tex(
    records: list[SourceRecord], bib: dict[str, dict], digest: str, path: Path
) -> None:
    lines = [
        "% Diese Datei wird von abgabe_korpus.py erzeugt. Nicht manuell bearbeiten.",
        f"% Korpus-Pruefsumme: {digest}",
        r"\section*{Abgeleiteter Quellenkorpus}",
        r"\begin{description}",
    ]
    for record in records:
        if not record.counted:
            continue
        lines.append(
            rf"  \item[{record.frage}, {record.status}, \cite{{{record.key}}}] "
            rf"\fullcite{{{record.key}}}. Relevanter Seitenbereich: "
            rf"S.\,{tex_escape_text(record.seiten)} "
            rf"({record.seitenzahl}\,S.)"
        )
    lines.extend([r"\end{description}", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_page_set(expr: str) -> set[int] | None:
    """Best-effort parser for overlap auditing; explicit count remains authoritative."""
    cleaned = (
        expr.replace(r"\,", " ")
        .replace("S.", "")
        .replace(" ", "")
        .replace(";", ",")
    )
    if not cleaned or re.search(r"[A-Za-z]", cleaned):
        return None
    pages: set[int] = set()
    for part in cleaned.split(","):
        if not part:
            continue
        m = re.fullmatch(r"([0-9]+)--([0-9]+)", part)
        if m:
            start, end = int(m.group(1)), int(m.group(2))
            if end < start:
                return None
            pages.update(range(start, end + 1))
            continue
        m = re.fullmatch(r"([0-9]+)", part)
        if m:
            pages.add(int(m.group(1)))
            continue
        return None
    return pages


def overlap_notes(records: list[SourceRecord]) -> list[str]:
    notes: list[str] = []
    by_key: dict[str, list[SourceRecord]] = {}
    for record in records:
        if record.counted:
            by_key.setdefault(record.key, []).append(record)
    for key, key_records in sorted(by_key.items()):
        if len(key_records) < 2:
            continue
        for i, left in enumerate(key_records):
            left_pages = parse_page_set(left.seiten)
            if left_pages is None:
                continue
            for right in key_records[i + 1 :]:
                right_pages = parse_page_set(right.seiten)
                if right_pages is None:
                    continue
                overlap = left_pages & right_pages
                if overlap:
                    notes.append(
                        f"- `{key}`: {left.frage} `{left.seiten}` und "
                        f"{right.frage} `{right.seiten}` ueberlappen auf "
                        f"{len(overlap)} parsebaren Seite(n)."
                    )
    return notes


def duplicate_key_notes(records: list[SourceRecord]) -> list[str]:
    by_key: dict[str, list[SourceRecord]] = {}
    for record in records:
        if record.counted:
            by_key.setdefault(record.key, []).append(record)
    notes: list[str] = []
    for key, key_records in sorted(by_key.items()):
        if len(key_records) > 1:
            refs = ", ".join(
                f"{rec.frage}/{rec.status} {rec.seiten} ({rec.seitenzahl} S.)"
                for rec in key_records
            )
            notes.append(f"- `{key}`: {refs}")
    return notes


def write_audit(
    records: list[SourceRecord],
    totals: dict[str, int],
    digest: str,
    tex_path: Path,
    bib_path: Path,
    path: Path,
) -> None:
    counted = [record for record in records if record.counted]
    uncounted = [record for record in records if not record.counted]
    duplicates = duplicate_key_notes(records)
    overlaps = overlap_notes(records)
    lines = [
        "# Abgabe-Korpus Audit",
        "",
        f"- TeX-Quelle: `{tex_path}`",
        f"- Bib-Quelle: `{bib_path}`",
        f"- Korpus-Pruefsumme: `{digest}`",
        f"- Gezaehlte Eintraege: {len(counted)}",
        f"- Nicht gezaehlte Eintraege: {len(uncounted)}",
        "",
        "## Seitenbudget",
        "",
        "| Frage | Seiten |",
        "|---|---:|",
    ]
    for frage in QUESTION_LABELS:
        lines.append(f"| {frage} | {totals[frage]} |")
    lines.append(f"| **Summe fragespezifisch** | **{sum(totals.values())}** |")
    lines.extend(["", "## Mehrfach verwendete BibKeys", ""])
    lines.extend(duplicates or ["Keine mehrfach verwendeten gezaehlten BibKeys."])
    lines.extend(["", "## Parsebare Seitenueberlappungen", ""])
    lines.extend(overlaps or ["Keine parsebaren Seitenueberlappungen gefunden."])
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def run(args: argparse.Namespace) -> int:
    root = Path.cwd()
    tex_path = (root / args.tex).resolve()
    bib_path = (root / args.bib).resolve()
    out_base = (root / args.out).resolve().with_suffix("")
    build_dir = (root / args.build_dir).resolve()
    build_dir.mkdir(parents=True, exist_ok=True)

    if not tex_path.exists():
        raise KorpusError(f"TeX-Datei nicht gefunden: {tex_path}")
    if not bib_path.exists():
        raise KorpusError(f"Bib-Datei nicht gefunden: {bib_path}")

    tex_text = collect_tex(tex_path)
    bib = load_bib(bib_path)
    records = parse_sources(tex_text)
    validate_records(records, bib, tex_text)

    overlaps = overlap_notes(records)
    if args.strict_overlap and overlaps:
        raise KorpusError(
            "Gezaehlte Seitenueberlappungen gefunden:\n" + "\n".join(overlaps)
        )

    digest = corpus_hash(records, bib)
    totals = totals_by_question(records)

    rows = [row_for(record, bib, digest) for record in records]
    rows_counted = [row for row in rows if row["Gezaehlt"] == "ja"]
    rows_uncounted = [row for row in rows if row["Gezaehlt"] == "nein"]

    csv_path = out_base.with_suffix(".csv")
    xlsx_path = out_base.with_suffix(".xlsx")

    enforce_existing_xlsx_matches(xlsx_path, records, args.force)
    write_csv(rows, csv_path)
    xlsx_written = write_xlsx(
        rows_counted, rows_uncounted, totals, digest, xlsx_path, tex_path, bib_path
    )
    write_budget_tex(totals, digest, build_dir / "abgabe_seitenbudget.tex")
    write_quellenliste_tex(records, bib, digest, build_dir / "abgabe_quellenliste.tex")
    write_audit(
        records,
        totals,
        digest,
        tex_path,
        bib_path,
        build_dir / "abgabe_korpus_audit.md",
    )

    print(f"[OK] {len(records)} Korpuseintraege validiert.")
    print(f"[OK] Seitenbudget: {', '.join(f'{k}={v}' for k, v in totals.items())}")
    print(f"[OK] Summe fragespezifisch: {sum(totals.values())}")
    print(f"[OK] Korpus-Pruefsumme: {digest}")
    print(f"[OK] CSV: {csv_path}")
    if xlsx_written:
        print(f"[OK] XLSX: {xlsx_path}")
    print(f"[OK] LaTeX-Seitenbudget: {build_dir / 'abgabe_seitenbudget.tex'}")
    print(f"[OK] Audit: {build_dir / 'abgabe_korpus_audit.md'}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate Abgabe_Quellen and LaTeX budget from \\AbgabeQuelle markers."
    )
    parser.add_argument("--tex", default="mpv.tex", help="TeX entry file (default: mpv.tex)")
    parser.add_argument("--bib", default="Quellen.bib", help="BibLaTeX file")
    parser.add_argument("--out", default="Abgabe_Quellen", help="Output basename")
    parser.add_argument("--build-dir", default="build", help="Generated TeX/audit directory")
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite an existing generated XLSX even if its rows differ.",
    )
    parser.add_argument(
        "--strict-overlap",
        action="store_true",
        help="Fail if the same BibKey has parseable overlapping counted page ranges.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return run(args)
    except KorpusError as exc:
        print("[FEHLER] Abgabe-Korpus ungueltig:", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
