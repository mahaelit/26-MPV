"""Analyse-Skript: Bestandsaufnahme vor der Zitat-Verifikation.

Liest:
  - Quellen.bib         -> alle BibKeys im Projekt
  - mpv.tex             -> welche Keys werden wie oft zitiert
  - Quellen v0.3.xlsx   -> welche Swisscovery-URLs / Statuswerte sind bereits
                           vom Nutzer gepflegt
  - MPV/Literatur/*.pdf -> welche PDFs liegen vor
  - MPV/26-MPV/Literatur/<key>/ -> welche Zielordner existieren

Gibt eine Uebersicht aus, die fuer den Verifikations-Plan gebraucht wird.
Diese Datei ist temporaer und wird nach dem Lauf wieder geloescht.
"""
from __future__ import annotations
import re
import sys
from collections import Counter
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent                                          # MPV/
EXTERNAL_LIT = ROOT / "Literatur"                           # beschaffte PDFs
INTERNAL_LIT = HERE / "Literatur"                           # Zielordner
BIB_PATH     = HERE / "Quellen.bib"
TEX_PATH     = HERE / "mpv.tex"
XLSX_PATH    = HERE / "Quellen v0.3.xlsx"


# ------------- 1) BibKeys aus Quellen.bib ---------------------------------
bib_text = BIB_PATH.read_text(encoding="utf-8")
BIB_ENTRY_RE = re.compile(r"@(\w+)\s*\{\s*([^,\s]+)\s*,", re.IGNORECASE)
bib_keys = [m.group(2) for m in BIB_ENTRY_RE.finditer(bib_text)]
bib_set  = set(bib_keys)
print(f"[1] Quellen.bib enthaelt {len(bib_keys)} Eintraege.\n")


# ------------- 2) cite-Keys aus mpv.tex -----------------------------------
tex_raw = TEX_PATH.read_text(encoding="utf-8")
# \verb|...|-Segmente und Kommentare ausblenden, damit das Literal
# "\cite{...}" in einem Hinweistext nicht als echtes Zitat gewertet wird.
tex = re.sub(r"\\verb[|@~!+#](?:.*?)[|@~!+#]", "", tex_raw)
tex = re.sub(r"(?m)(?<!\\)%.*$", "", tex)
CITE_RE = re.compile(
    r"\\(?:cite|textcite|parencite|autocite|footcite|citep|citet|"
    r"citealp|citeauthor|citeyear)[a-zA-Z]*"
    r"(?:\[[^\]]*\]){0,2}\{([^}]+)\}"
)
cited = Counter()
for m in CITE_RE.finditer(tex):
    for k in m.group(1).split(","):
        k = k.strip()
        if k:
            cited[k] += 1

print(f"[2] mpv.tex  enthaelt {sum(cited.values())} cite-Aufrufe "
      f"auf {len(cited)} unterschiedliche BibKeys.")

not_cited   = sorted(bib_set - set(cited))
not_in_bib  = sorted(set(cited) - bib_set)
print(f"    - nie zitiert ({len(not_cited)}): {', '.join(not_cited) or '—'}")
if not_in_bib:
    print(f"    - im TeX zitiert, aber NICHT in .bib ({len(not_in_bib)}): "
          f"{', '.join(not_in_bib)}")
    # Zeilen im TeX zeigen
    for miss in not_in_bib:
        for li, line in enumerate(tex.splitlines(), start=1):
            if miss in line:
                print(f"      Zeile {li}: {line.strip()[:140]}")
print()


# ------------- 3) XLSX v0.3 auslesen --------------------------------------
xlsx_info = {}
if XLSX_PATH.exists():
    try:
        from openpyxl import load_workbook
        wb = load_workbook(XLSX_PATH, data_only=True)
        ws = wb["Quellen"] if "Quellen" in wb.sheetnames else wb.active
        h = [c.value for c in ws[1]]
        idx = {n: i for i, n in enumerate(h)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[idx["BibKey"]] if "BibKey" in idx else None
            if not key:
                continue
            xlsx_info[str(key)] = {
                "status":     row[idx["Status"]] if "Status" in idx else None,
                "swisscov":   row[idx["Swisscovery_URL"]] if "Swisscovery_URL" in idx else None,
                "pfad":       row[idx["Pfad_lokal"]] if "Pfad_lokal" in idx else None,
                "beschaffung": row[idx["Beschaffung"]] if "Beschaffung" in idx else None,
                "bemerkung":  row[idx["Bemerkung"]] if "Bemerkung" in idx else None,
            }
        print(f"[3] '{XLSX_PATH.name}' enthaelt {len(xlsx_info)} Zeilen.")
    except Exception as exc:
        print(f"[3] XLSX-Lesefehler: {exc}")
else:
    print(f"[3] '{XLSX_PATH.name}' NICHT gefunden.")

swisscov_ok = sum(1 for v in xlsx_info.values() if v["swisscov"])
print(f"    - Swisscovery_URL bereits gepflegt: {swisscov_ok}/{len(xlsx_info)}")
print()


# ------------- 4) PDFs im externen Literatur-Ordner -----------------------
pdfs = sorted(p for p in EXTERNAL_LIT.iterdir() if p.is_file() and p.suffix.lower() in (".pdf", ".epub"))
print(f"[4] {len(pdfs)} PDF/EPUB-Dateien in {EXTERNAL_LIT}:\n")
for p in pdfs:
    size_mb = p.stat().st_size / (1024 * 1024)
    print(f"    {size_mb:6.2f} MB  {p.name}")
print()


# ------------- 5) Zielordner 26-MPV/Literatur -----------------------------
targets = sorted(d for d in INTERNAL_LIT.iterdir() if d.is_dir()) if INTERNAL_LIT.exists() else []
has_source_pdf = [d for d in targets if (d / "source.pdf").exists()]
print(f"[5] {len(targets)} Zielordner in {INTERNAL_LIT}, davon {len(has_source_pdf)} mit source.pdf.")
print()


# ------------- 6) Vorschlag: PDF -> BibKey Mapping (heuristisch) ----------
def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9äöüß]+", "", s.lower())

# Metadaten pro BibKey aus Quellen.bib (Autor-Nachname + Jahr + Titelwoerter)
BIB_FIELD_RE = re.compile(
    r"@(\w+)\s*\{\s*([^,]+),(.*?)\n\}\n",
    re.IGNORECASE | re.DOTALL,
)
bib_meta = {}
for m in BIB_FIELD_RE.finditer(bib_text):
    key  = m.group(2).strip()
    body = m.group(3)
    def field(name):
        m2 = re.search(rf"\b{name}\s*=\s*\{{(.*?)\}},", body, re.IGNORECASE | re.DOTALL)
        return m2.group(1).strip() if m2 else ""
    author = field("author") or field("editor")
    surname = re.split(r",| and ", author)[0].strip().strip("{}")
    year   = field("year")
    title  = field("title")
    bib_meta[key] = (norm(surname), year, norm(title))

# Nur fuer Anzeige: erraten welchem BibKey jedes PDF am naechsten kommt.
print("[6] Heuristische Zuordnung PDF -> BibKey (Vorschlag, MUSS vom Nutzer bestaetigt werden):")
print()
for p in pdfs:
    fname_norm = norm(p.stem)
    best = None
    best_score = 0
    for k, (surname, year, title_n) in bib_meta.items():
        score = 0
        if surname and surname in fname_norm:
            score += 3
        if year and year in fname_norm:
            score += 2
        # Titel-Teil-Match (erste Woerter ab 5 Zeichen)
        title_words = [w for w in re.findall(r"[a-z0-9äöüß]{5,}", title_n)][:4]
        score += sum(1 for w in title_words if w in fname_norm)
        # Key-Bestandteile
        key_words = [w for w in re.findall(r"[a-z0-9]{4,}", k.lower())]
        score += sum(1 for w in key_words if w in fname_norm)
        if score > best_score:
            best_score, best = score, k
    mark = "??" if best_score < 3 else "ok" if best_score >= 5 else "? "
    print(f"    [{mark}|{best_score:2}] {p.name[:72]:<72} -> {best or '—'}")


# ------------- 7) Priorisierung: BibKeys nach Zitationen -------------------
print()
print("[7] BibKeys nach Zitationshaeufigkeit (Verifikations-Prioritaet):")
print()
# Set der beschafften PDFs (heuristische Zuordnung fuer Anzeige)
pdf_assigned = set()
for p in pdfs:
    fname_norm = norm(p.stem)
    best = None; best_score = 0
    for k, (surname, year, title_n) in bib_meta.items():
        score = 0
        if surname and surname in fname_norm: score += 3
        if year and year in fname_norm: score += 2
        title_words = [w for w in re.findall(r"[a-z0-9äöüß]{5,}", title_n)][:4]
        score += sum(1 for w in title_words if w in fname_norm)
        key_words = [w for w in re.findall(r"[a-z0-9]{4,}", k.lower())]
        score += sum(1 for w in key_words if w in fname_norm)
        if score > best_score: best_score, best = score, k
    if best and best_score >= 5:
        pdf_assigned.add(best)

xlsx_swisscov = {k for k, v in xlsx_info.items() if v.get("swisscov")}

print(f"{'BibKey':35} | Zit | PDF | SC-URL")
print("-"*70)
for k, n in sorted(cited.items(), key=lambda x: (-x[1], x[0])):
    pdf_flag = "JA" if k in pdf_assigned else "—"
    sc_flag  = "JA" if k in xlsx_swisscov else "—"
    print(f"{k:35} | {n:3} | {pdf_flag:3} | {sc_flag}")
